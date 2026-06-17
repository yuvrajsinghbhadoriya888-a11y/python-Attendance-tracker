import openpyxl
from datetime import datetime

file = openpyxl.load_workbook("book2.xlsx")
sheet = file.active


while True:
    print("[Y] to Mark attendance")
    print("[N] to Exit")

    choice = input("Enter your choice: ").upper()

    if choice == "Y":
        pass
    elif choice == "N":
        print("Exiting the program.\nGood Bye!")
        exit()
    else:
        print("Invalid choice. Please select Y or N.")


    roll_number = int(input("Enter roll number: "))

    row_number = None

    for r, row in enumerate(sheet.iter_rows(min_row=2), start=2):
     if row[0].value == roll_number:
        row_number = r
        break

    if row_number is None:
     print("Roll number not found.")
     continue

    print("Student found at row:", row_number)


    attendance = input("Enter attendance (P/A): ").upper()
    while attendance not in ["P", "A"]:
     attendance = input("Enter attendance (P/A): ").upper()


    today = datetime.today().date()
    date_column = None

    for cell in sheet[1]:
     if isinstance(cell.value, datetime) and cell.value.date() == today:
        date_column = cell.column
        break

    if date_column is None:
     date_column = sheet.max_column + 1
     sheet.cell(row=1, column=date_column).value = today


    sheet.cell(row=row_number, column=date_column).value = attendance

    file.save("book2.xlsx")
    print("Attendance marked successfully!")






    